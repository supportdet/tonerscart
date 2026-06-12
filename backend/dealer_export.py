"""Admin → Dealer Profile ZIP export.

Bundles a Dealer Profile PDF (reportlab), an Analytics Excel (openpyxl), and
every KYC document the supplier uploaded into a single in-memory ZIP that is
streamed to the admin browser. Used by the "Download Full Profile" button on
the admin dealer profile page (`/admin/dealers/:id`).
"""
# ruff: noqa: F403, F405
from __future__ import annotations

import io
import logging
import re
import zipfile
from datetime import datetime
from typing import Any, Dict, List, Optional

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import mm
from reportlab.platypus import (
    SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle, PageBreak,
)

logger = logging.getLogger("tonerscart")


DOC_LABELS = {
    "doc_id_proof": "ID Proof (Aadhaar/Passport)",
    "doc_bank_proof": "Cancelled Cheque / Bank Proof",
    "doc_gst": "GST Certificate",
    "doc_pan": "PAN Card",
    "doc_brand_authorization": "Brand Authorization",
    "doc_shop_photo": "Shop Photo",
    "doc_address_proof": "Address Proof",
}

_RUPEE = "Rs. "  # reportlab default font does not include the ₹ glyph reliably


def _safe_filename(name: str) -> str:
    s = re.sub(r"[^A-Za-z0-9._-]+", "_", (name or "dealer").strip())
    return s[:80] or "dealer"


def _fmt_date(d: Optional[str]) -> str:
    if not d:
        return "—"
    try:
        return datetime.fromisoformat(str(d).replace("Z", "+00:00")).strftime("%d %b %Y")
    except Exception:
        return str(d)[:10]


def _fmt_money(n: Optional[float]) -> str:
    try:
        return f"{_RUPEE}{int(round(float(n or 0))):,}"
    except Exception:
        return f"{_RUPEE}0"


def _build_profile_pdf(supplier: Dict[str, Any], stats: Dict[str, Any],
                       documents: Dict[str, Any]) -> bytes:
    """Returns a single-PDF byte string with business / bank / KYC sections."""
    buf = io.BytesIO()
    doc = SimpleDocTemplate(
        buf, pagesize=A4,
        leftMargin=18 * mm, rightMargin=18 * mm,
        topMargin=18 * mm, bottomMargin=18 * mm,
        title=f"{supplier.get('business_name') or 'Dealer'} — TonersCart Dealer Profile",
    )
    styles = getSampleStyleSheet()
    h1 = ParagraphStyle("h1", parent=styles["Heading1"], fontSize=18, spaceAfter=4, textColor=colors.HexColor("#0A0A0B"))
    sub = ParagraphStyle("sub", parent=styles["Normal"], fontSize=10, textColor=colors.HexColor("#6E6E73"), spaceAfter=14)
    h2 = ParagraphStyle("h2", parent=styles["Heading2"], fontSize=13, spaceBefore=14, spaceAfter=6, textColor=colors.HexColor("#0A0A0B"))
    body = ParagraphStyle("body", parent=styles["Normal"], fontSize=10, leading=14, textColor=colors.HexColor("#1D1D1F"))

    story: List[Any] = []
    story.append(Paragraph(f"{supplier.get('business_name') or 'Dealer'} — TonersCart Dealer Profile", h1))
    story.append(Paragraph(
        f"Seller ID <b>{supplier.get('seller_id') or '—'}</b> · Generated on {datetime.now().strftime('%d %b %Y, %I:%M %p')}",
        sub,
    ))

    def _kv_table(rows: List[List[str]]) -> Table:
        tbl = Table(rows, colWidths=[55 * mm, 110 * mm])
        tbl.setStyle(TableStyle([
            ("BACKGROUND", (0, 0), (0, -1), colors.HexColor("#F5F5F7")),
            ("TEXTCOLOR", (0, 0), (0, -1), colors.HexColor("#6E6E73")),
            ("FONTNAME", (0, 0), (0, -1), "Helvetica-Bold"),
            ("FONTSIZE", (0, 0), (-1, -1), 9.5),
            ("VALIGN", (0, 0), (-1, -1), "TOP"),
            ("LEFTPADDING", (0, 0), (-1, -1), 8),
            ("RIGHTPADDING", (0, 0), (-1, -1), 8),
            ("TOPPADDING", (0, 0), (-1, -1), 5),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 5),
            ("LINEBELOW", (0, 0), (-1, -1), 0.25, colors.HexColor("#E5E5EA")),
        ]))
        return tbl

    # Business Details
    story.append(Paragraph("Business Details", h2))
    story.append(_kv_table([
        ["Company name", supplier.get("business_name") or "—"],
        ["Seller ID", supplier.get("seller_id") or "—"],
        ["Owner / contact", supplier.get("contact_person") or "—"],
        ["GST number", supplier.get("gst_number") or "—"],
        ["PAN number", supplier.get("pan_number") or "—"],
        ["Business address", supplier.get("business_address") or "—"],
        ["City", supplier.get("city") or "—"],
        ["State / Pincode", " · ".join(filter(None, [supplier.get("state"), supplier.get("pincode")])) or "—"],
        ["Phone", supplier.get("phone") or "—"],
        ["Email", supplier.get("account_email") or supplier.get("email") or "—"],
        ["Registration date", _fmt_date(supplier.get("registration_date") or supplier.get("created_at"))],
        ["Approval date", _fmt_date(supplier.get("approved_at"))],
        ["Verification status", "Suspended" if supplier.get("is_suspended") else "Active"],
    ]))

    # Bank Details
    story.append(Paragraph("Bank Details", h2))
    story.append(_kv_table([
        ["Account holder name", supplier.get("account_holder_name") or "—"],
        ["Account number", supplier.get("account_number") or "—"],
        ["IFSC code", supplier.get("ifsc_code") or "—"],
        ["Bank name", supplier.get("bank_name") or "—"],
        ["Branch", supplier.get("bank_branch") or "—"],
    ]))

    # KYC documents
    story.append(Paragraph("KYC Documents", h2))
    doc_rows = []
    for key, label in DOC_LABELS.items():
        if supplier.get(key) or (documents or {}).get(key):
            doc_rows.append([label, _fmt_date(supplier.get("submitted_at") or supplier.get("created_at"))])
    if not doc_rows:
        story.append(Paragraph("No KYC documents on file.", body))
    else:
        tbl = Table([["Document", "Submitted on"]] + doc_rows, colWidths=[110 * mm, 55 * mm])
        tbl.setStyle(TableStyle([
            ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#0A0A0B")),
            ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
            ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
            ("FONTSIZE", (0, 0), (-1, -1), 9.5),
            ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
            ("LEFTPADDING", (0, 0), (-1, -1), 8),
            ("RIGHTPADDING", (0, 0), (-1, -1), 8),
            ("TOPPADDING", (0, 0), (-1, -1), 5),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 5),
            ("LINEBELOW", (0, 0), (-1, -1), 0.25, colors.HexColor("#E5E5EA")),
        ]))
        story.append(tbl)

    # Footer note
    story.append(Spacer(1, 14))
    story.append(Paragraph(
        "Confidential — internal admin export. Bank and KYC data must be handled per TonersCart's privacy policy.",
        ParagraphStyle("note", parent=body, fontSize=8.5, textColor=colors.HexColor("#86868B")),
    ))

    doc.build(story)
    return buf.getvalue()


def _build_analytics_xlsx(supplier: Dict[str, Any], stats: Dict[str, Any],
                          orders: List[Dict[str, Any]]) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "Dealer Analytics"

    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="0A0A0B")
    label_font = Font(bold=True, color="0A0A0B")
    money_fmt = '"\u20B9"#,##0'

    ws["A1"] = "Metric"
    ws["B1"] = "Value"
    for c in ("A1", "B1"):
        ws[c].font = header_font
        ws[c].fill = header_fill
        ws[c].alignment = Alignment(horizontal="left")

    # Status buckets
    status_buckets = {"requested": 0, "accepted": 0, "shipped": 0, "delivered": 0, "completed": 0, "cancelled": 0}
    total_payout = 0.0
    total_commission = 0.0
    for o in orders or []:
        s = (o.get("status") or "").lower()
        if s in status_buckets:
            status_buckets[s] += 1
        total = float(o.get("total") or 0)
        commission = float(o.get("commission_amount") or 0)
        # Best-effort payout = total - commission - delivery_charge (charged to buyer; dealer keeps net)
        if (o.get("status") or "").lower() in ("completed", "delivered"):
            total_payout += max(0.0, total - commission)
        total_commission += commission

    order_count = stats.get("order_count") or len(orders or [])
    gmv = float(stats.get("gmv") or 0)
    aov = (gmv / order_count) if order_count else 0
    last_active = None
    for o in orders or []:
        ts = o.get("created_at")
        if ts and (last_active is None or ts > last_active):
            last_active = ts

    rows = [
        ("Total Orders", order_count, None),
        ("Total GMV", gmv, money_fmt),
        ("Total Commission Paid to TonersCart", round(total_commission or float(stats.get("commission_earned") or 0), 2), money_fmt),
        ("Total Payouts to Dealer", round(total_payout, 2), money_fmt),
        ("Orders — Requested", status_buckets["requested"], None),
        ("Orders — Confirmed (Accepted)", status_buckets["accepted"], None),
        ("Orders — Dispatched (Shipped)", status_buckets["shipped"], None),
        ("Orders — Delivered", status_buckets["delivered"], None),
        ("Orders — Completed", status_buckets["completed"], None),
        ("Orders — Cancelled", status_buckets["cancelled"], None),
        ("Listings — Toners", stats.get("toner_count") or 0, None),
        ("Listings — Printers", stats.get("printer_count") or 0, None),
        ("Listings — Papers", stats.get("paper_count") or 0, None),
        ("Listings — Scanners", stats.get("scanner_count") or 0, None),
        ("Listings — Consumables", stats.get("consumable_count") or 0, None),
        ("Total Listings", stats.get("listing_count") or 0, None),
        ("Average Order Value", round(aov, 2), money_fmt),
        ("Member Since", _fmt_date(supplier.get("registration_date") or supplier.get("created_at")), None),
        ("Last Active", _fmt_date(last_active), None),
    ]
    for i, (label, value, fmt) in enumerate(rows, start=2):
        ws.cell(row=i, column=1, value=label).font = label_font
        c = ws.cell(row=i, column=2, value=value)
        if fmt:
            c.number_format = fmt

    ws.column_dimensions["A"].width = 42
    ws.column_dimensions["B"].width = 28

    out = io.BytesIO()
    wb.save(out)
    return out.getvalue()


def build_dealer_export_zip(sb_admin, supplier_id: str, detail_payload: Dict[str, Any]) -> tuple[bytes, str]:
    """Build the ZIP in memory and return (bytes, filename)."""
    supplier = detail_payload.get("supplier") or {}
    stats = detail_payload.get("stats") or {}
    documents = detail_payload.get("documents") or {}
    orders = detail_payload.get("orders") or []

    safe = _safe_filename(supplier.get("business_name") or "dealer")
    zip_filename = f"{safe}__TonersCart_Profile.zip"

    pdf_bytes = _build_profile_pdf(supplier, stats, documents)
    xlsx_bytes = _build_analytics_xlsx(supplier, stats, orders)

    # Pull each KYC document fresh from Supabase Storage (admin service key bypasses RLS).
    doc_files: List[tuple[str, bytes]] = []
    doc_source = dict(supplier)
    # Merge any paths still living on the original suppliers_pending row.
    if supplier.get("user_id"):
        try:
            pend = sb_admin.table("suppliers_pending").select("*").eq("user_id", supplier["user_id"]).maybe_single().execute()
            if pend and pend.data:
                for f in DOC_LABELS.keys():
                    if not doc_source.get(f) and pend.data.get(f):
                        doc_source[f] = pend.data[f]
        except Exception:
            pass

    for field, label in DOC_LABELS.items():
        path = doc_source.get(field)
        if not path:
            continue
        try:
            blob = sb_admin.storage.from_("supplier-documents").download(path)
            if not blob:
                continue
            # Preserve the original extension if present.
            ext = path.rsplit(".", 1)[-1].lower() if "." in path.rsplit("/", 1)[-1] else "bin"
            doc_files.append((f"KYC/{_safe_filename(label)}.{ext}", blob))
        except Exception as e:
            logger.warning("dealer export: skip doc %s for %s: %s", field, supplier_id, e)

    buf = io.BytesIO()
    with zipfile.ZipFile(buf, "w", zipfile.ZIP_DEFLATED) as zf:
        zf.writestr(f"{safe}__Dealer_Profile.pdf", pdf_bytes)
        zf.writestr("Dealer_Analytics.xlsx", xlsx_bytes)
        for name, blob in doc_files:
            zf.writestr(name, blob)
        zf.writestr("README.txt", (
            f"TonersCart — Dealer Profile Export\n"
            f"Generated: {datetime.now().strftime('%d %b %Y, %I:%M %p')}\n"
            f"Dealer: {supplier.get('business_name') or '—'}\n"
            f"Seller ID: {supplier.get('seller_id') or '—'}\n\n"
            f"Files included:\n"
            f"  • {safe}__Dealer_Profile.pdf — Business, Bank & KYC summary\n"
            f"  • Dealer_Analytics.xlsx — Orders, GMV, Commission, Payouts, Listings\n"
            f"  • KYC/ — Original uploaded documents (GST, PAN, bank proof, etc.)\n"
        ))
    return buf.getvalue(), zip_filename
